import openpyxl
import npx
import numpy as np
import json
import yaml

wb = openpyxl.load_workbook("Leeds.xlsx")

ws = wb.worksheets[0]
dv = [ws[f"A{k}"].value for k in range(3, 310)]

white = [ws["B3"].value, ws["C3"].value, ws["D3"].value]

pairs = [
    [
        [ws[f"E{k}"].value, ws[f"F{k}"].value, ws[f"G{k}"].value],
        [ws[f"H{k}"].value, ws[f"I{k}"].value, ws[f"J{k}"].value],
    ]
    for k in range(3, 310)
]

# find duplicates in the pairs
pairs = np.array(pairs)
pairs = pairs.reshape(-1, 3)
uniques, idx = npx.unique_rows(pairs, return_inverse=True)
idx = idx.reshape((-1, 2))

d = {"reference_white": white, "dv": dv, "pairs": idx.tolist(), "xyz": uniques.tolist()}

with open("leeds.yaml", "w") as f:
    yaml.dump(d, f)

with open("leeds.json", "w") as f:
    json.dump(d, f, indent=2)
