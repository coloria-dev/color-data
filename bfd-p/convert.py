import json

import numpy as np
import openpyxl
import yaml

wb = openpyxl.load_workbook("BFD-P.xlsx")
ws = wb.worksheets[0]

ranges = {
    "bfd-d65": (3, 2031),
    "bfd-c": (2031, 2231),
    "bfd-m": (2231, 2779),
}

for name, (k0, k1) in ranges.items():
    dv = [ws[f"A{k}"].value for k in range(k0, k1)]

    reference_white = [ws[f"B{k0}"].value, ws[f"C{k0}"].value, ws[f"D{k0}"].value]

    # The first 2028 are BFD-D65
    pairs = [
        [
            [ws[f"E{k}"].value, ws[f"F{k}"].value, ws[f"G{k}"].value],
            [ws[f"H{k}"].value, ws[f"I{k}"].value, ws[f"J{k}"].value],
        ]
        for k in range(k0, k1)
    ]

    # find duplicates in the pairs
    pairs = np.array(pairs)
    pairs = pairs.reshape(-1, 3)
    uniques, idx = np.unique(pairs, return_inverse=True, axis=0)
    idx = idx.reshape((-1, 2))

    d = {
        "reference_white": reference_white,
        "dv": dv,
        "pairs": idx.tolist(),
        "xyz": uniques.tolist(),
    }

    with open(f"{name}.yaml", "w") as f:
        yaml.dump(d, f)

    with open(f"{name}.json", "w") as f:
        json.dump(d, f)
