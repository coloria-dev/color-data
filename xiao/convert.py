"""
Helper tool for converting XLSX data to YAML.
"""
import numpy as np
import openpyxl
import yaml

lst = [(0, "red"), (1, "green"), (2, "yellow"), (3, "blue")]
averages = {}

wb = openpyxl.load_workbook("xiao.xlsx")
for index, color in lst:
    filename = f"unique_{color}.yaml"
    # first worksheet
    ws = wb.worksheets[index]
    session1_x = [ws[f"C{k}"].value for k in range(4, 1669)]
    session1_y = [ws[f"D{k}"].value for k in range(4, 1669)]
    session1_z = [ws[f"E{k}"].value for k in range(4, 1669)]

    session2_x = [ws[f"G{k}"].value for k in range(4, 1669)]
    session2_y = [ws[f"H{k}"].value for k in range(4, 1669)]
    session2_z = [ws[f"I{k}"].value for k in range(4, 1669)]

    session3_x = [ws[f"K{k}"].value for k in range(4, 1669)]
    session3_y = [ws[f"L{k}"].value for k in range(4, 1669)]
    session3_z = [ws[f"M{k}"].value for k in range(4, 1669)]

    data = np.array(
        [
            [session1_x, session2_x, session3_x],
            [session1_y, session2_y, session3_y],
            [session1_z, session2_z, session3_z],
        ]
    ).T
    data = data.reshape(-1, 9, 3, 3)
    data = np.moveaxis(data, 1, 2)

    averages[color] = np.average(data, axis=(0, 1)).tolist()

    with open(filename, "w") as f:
        yaml.dump(data.tolist(), f)


with open("averages.yaml", "w") as f:
    yaml.dump(averages, f)
